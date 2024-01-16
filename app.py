import openpyxl
from openpyxl import load_workbook
import os ,time
import pywss
import uuid
import re
import pymysql
from urllib.parse import quote_plus as urlquote
from sqlalchemy import create_engine
import pandas as pd
import json
import zlib
import threading
from urllib.parse import unquote
from flask import Flask, render_template, request
import signal  
 
# app = Flask(__name__, template_folder='.',static_folder="",static_url_path="")
 
class Pool:
    lock = threading.Lock()
    pool = {}

    @classmethod
    def add(cls, uid, ctx):
        with cls.lock:
            cls.pool[uid] = ctx

    @classmethod
    def delete(cls, uid):
        with cls.lock:
            cls.pool.pop(uid, None)

    @classmethod
    def notify(cls, data, by):
        with cls.lock:
            for uid, ctx in cls.pool.items():  # type: pywss.Context
                if uid == by:
                    continue
                ctx.ws_write(data)

def handle_sigint(signal, frame):  
    # 清理操作，例如关闭数据库连接、保存数据等  
    print('Flask app is shutting down...')  
    app.stop()  
  
signal.signal(signal.SIGINT, handle_sigint)  

# @app.route('/')  

def DataToJson():
    
    conn = pymysql.Connect(host='127.0.0.1', port=3306, user='root', passwd='password', charset='utf8', db='unicom')
    #  设置自己的mysql信息
    cur = conn.cursor()
    sql = "select * from testlog"  #  选择具体的数据库db='unicom'下的testlog
    cur.execute(sql)
    data = cur.fetchall()
    cur.close()
    conn.close()
    jsonData = []
    for i, cell in enumerate(data):
        for j, value in enumerate(cell):
            result = {}
            result['r'] = int(i)
            result['c'] = int(j)
            result['v'] = str(value)
            jsonData.append(result)
    return (jsonData)
    #  luckysheet的要求导出格式
 
 
 
 
def load(ctx: pywss.Context):
    jsonData = DataToJson()
    data = json.dumps([
        {
            "name": "Cell",
            "index": "sheet_01",
            "order": 0,
            "status": 1,
            "celldata": jsonData,
        }])  # json.dumps将一个Python数据结构转换为JSON
    # json.dumps()用于将dict类型的数据转成str
    ctx.write(data)  # 写入json的数据

def excel_to_mysql():
    userName = "root"
    password = "password"
    dbHost = "127.0.0.1"
    dbPort = 3306
    dbName = "unicom"
    engine = create_engine(f'mysql+pymysql://{userName}:{urlquote(password)}@{dbHost}:{dbPort}/{dbName}?charset=utf8')
    data_frame = pd.read_excel('test.xlsx')
    data_frame = data_frame.where(data_frame.notnull(), '')
    data_frame.to_sql(name='testlog', con=engine, index=False, if_exists='replace')
 
def update(ctx: pywss.Context):
    # 升级 WebSocket
    err = pywss.WebSocketUpgrade(ctx)
    if err:
        ctx.log.error(err)
        ctx.set_status_code(pywss.StatusBadRequest)
        return
    uid = str(uuid.uuid4())
    Pool.add(uid, ctx)

    try:
        # 轮询获取消息
        while True:
            data = ctx.ws_read()
            if data == b"rub":  # 心跳检测
                continue
            data_raw = data.decode().encode('iso-8859-1')  # 转编码
            data_unzip = unquote(zlib.decompress(data_raw, 16).decode())  # 解压缩
            json_data = json.loads(data_unzip)
            resp_data = {
                "data": data_unzip,
                "id": uid,
                "returnMessage": "success",
                "status": 0,
                "type": 3,
                "username": uid,
            }
            if json_data.get("t") != "mv":
                resp_data["type"] = 2
            resp = json.dumps(resp_data).encode()
            Pool.notify(resp, uid)
    except:
        pass
    finally:
        ctx.log.warning(f"{uid} exit")
        Pool.delete(uid)

# @app.route('/test',methods=['GET','POST'])
def index(ctx: pywss.Context):
    if ctx.method == "POST":
        s = unquote(ctx.body())
        match=re.search(r'exportdata=',s)
        if not match:
            return
        b = json.loads(s[match.end():len(s)])
        # a = ctx.form().get('exportdata')
        # b = json.loads(a)
        for key, data in b.items():
            if key == 'celldata':
                list1 = []
                for i in data:
                    for j in i['v']:
                        if j == 'm':
                            row = i['r'] + 2
                            col = i['c'] + 1
                            value = i['v']['m']
                            list1.extend([[row, col, value]])
                #print(list1)
                if os.path.exists('test.xlsx') == False:
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    for i in list1:
                        sheet.cell(i[0], i[1]).value = i[2]
                    workbook.save('test.xlsx')
                    time.sleep(0.1)
                    excel_to_mysql()
                if os.path.exists('test.xlsx') == True:
                    os.remove('test.xlsx')
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    for i in list1:
                        sheet.cell(i[0], i[1]).value = i[2]
                    workbook.save('test.xlsx')
                    time.sleep(0.1)
                    excel_to_mysql()
 
# @app.route('/')  
def cool(ctx: pywss.Context):  
    # return render_template('index.html')
    # with open('index.html', 'r') as file:  
    # # 读取文件内容  
    #     content = file.read()
    ctx.write('cool')

# @app.route('/hi')  
def hi(ctx: pywss.Context):  
    print("--")
    ctx.write('Hello, World!' )
    # return 'Hello, World!' 


if __name__ == '__main__':
    app = pywss.App()
    # # 注册静态资源
    app.static("/static", ".")
    # # 注册 luckysheet 路由
    party = app.party("/luckysheet/api")
    party.post("/loadUrl", load)
    party.get("/updateUrl", update)

    # party = app.party("/test")
    app.get("/test", index)
    app.post("/test", index)


    # app.get("/", cool)
    app.get("/hi", hi)
    # 启动服务
    # app.run()
    app.run(host="127.0.0.1", port=5001)